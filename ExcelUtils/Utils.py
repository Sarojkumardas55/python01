import openpyxl
def getalldata(file,sheet):
# Replace 'your_file.xlsx' with the actual name of your Excel file
#file_path = 'D:\Saroj\data.xlsx'

# Load the workbook
  workbook = openpyxl.load_workbook(file)

# Choose the sheet you want to read from (replace 'Sheet1' with the actual sheet name)
  sheet = workbook[sheet]

# Get the number of rows and columns in the sheet
  max_row = sheet.max_row
  max_column = sheet.max_column

# Iterate through rows and columns to read data
  for i in range(1, max_row + 1):
    for j in range(1, max_column + 1):
        # Access cell value using cell coordinates (i, j)
        cell_value = sheet.cell(row=i, column=j).value
        print(cell_value, end='\t')  # Print the cell value (tab-separated)
    print()  # Move to the next row
  workbook.close()


def getrowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    max_row = sheet.max_row
    return max_row
def getCellCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    max_column = sheet.max_column
    return max_column
def getCellData(file,sheetName,row_num,cel_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    cell_value = sheet.cell(row=row_num, column=cel_num).value
    return cell_value
def writeCellData(file,sheetName,row_num,cel_num,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=row_num, column=cel_num).value=data
    workbook.save(file)
# Close the workbook when done

